import pandas as pd
from tqdm import tqdm
import matplotlib
matplotlib.use('TkAgg')
import time
import os
from openpyxl import load_workbook

from mdh_vfs import MDHVFS
import param_config

dataset = ['Change123']
data_path = "../SyntheticData/normal/"
result_path = 'result/normal/'
method_test = ['HD']
for method in method_test:
    params_list = param_config.config()
    model_params = param_config.create_param_combination(params_list)

    for k in range(len(model_params)):
        params = model_params[k]
        for data in dataset:
            path_data = f"{data_path}{data}.xlsx"
            df = pd.read_excel(path_data, sheet_name="All_Data")
            X = df.iloc[:, :-1].values
            Y = df.iloc[:, -1].values

            detector = MDHVFS(hist_window_size=params['window'],new_window_size=0.1*params['window'],
                                       referNum=params['referNum'], referValue=params['referValue'],
                                       warn=params['warningNum'], update_thre=params['updateThres'])

            num_sample = X.shape[0]
            instances_seen = 0
            uni_fea = []
            drift_results = []
            start_time = time.time()
            for t in tqdm(range(num_sample)):
                instances_seen += 1
                x_t = X[t, :]

                drift, drift_feature, drift_position = detector.detect(x_t)

                if drift:
                    print('driftPosition:', drift_position)
                    print('driftFeature:', drift_feature)

                    for feature in drift_feature:
                        drift_results.append({'drift_position': drift_position,
                                              'drift_feature': feature})

            end_time = time.time()
            runtime = end_time - start_time
            print('runtime: ', end_time - start_time)

            excel_file = f"{result_path}{data}_{method}_{params['referNum']}_{params['referValue']}.xlsx"
            sheet_name = f"{params['window']}_{params['referNum']}_{params['referValue']}_{params['warningNum']}" \
                         f"_{params['updateThres']}}"

            os.makedirs(result_path, exist_ok=True)

            if drift_results:
                df_result = pd.DataFrame(drift_results)

                if os.path.exists(excel_file):
                    book = load_workbook(excel_file)
                    if sheet_name in book.sheetnames:
                        del book[sheet_name]
                    book.save(excel_file)

                    with pd.ExcelWriter(excel_file, mode='a', engine='openpyxl') as writer:
                        df_result.to_excel(writer, sheet_name=sheet_name, index=False)
                else:
                    with pd.ExcelWriter(excel_file, mode='w', engine='openpyxl') as writer:
                        df_result.to_excel(writer, sheet_name=sheet_name, index=False)

                print(f"漂移检测结果已保存到 {excel_file} 的 {sheet_name} sheet中，共检测到 {len(drift_results)} 条漂移记录")
            else:
                print("未检测到任何漂移")

            runtime_data = {'方法名称': [sheet_name], 'runtime': [runtime]}
            df_runtime = pd.DataFrame(runtime_data)

            if os.path.exists(excel_file):
                try:
                    existing_runtime = pd.read_excel(excel_file, sheet_name='runtime')
                    if sheet_name in existing_runtime['方法名称'].values:
                        existing_runtime.loc[existing_runtime['方法名称'] == sheet_name, 'runtime'] = runtime
                        df_runtime = existing_runtime
                    else:
                        df_runtime = pd.concat([existing_runtime, df_runtime], ignore_index=True)
                except:
                    pass

            if os.path.exists(excel_file):
                with pd.ExcelWriter(excel_file, mode='a', if_sheet_exists='replace', engine='openpyxl') as writer:
                    df_runtime.to_excel(writer, sheet_name='runtime', index=False)
            else:
                with pd.ExcelWriter(excel_file, mode='w', engine='openpyxl') as writer:
                    df_runtime.to_excel(writer, sheet_name='runtime', index=False)